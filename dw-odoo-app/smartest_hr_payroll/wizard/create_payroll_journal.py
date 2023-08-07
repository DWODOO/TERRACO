# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
import base64
import os
import pdb
from datetime import datetime

from openpyxl import workbook, Workbook
from odoo import api, fields, models
from openpyxl.styles import Color, PatternFill, Font, Border, Side


class PayrollJournal(models.TransientModel):
    _name = 'create.payroll.journal'
    _description = 'Payroll journal'

    def _get_default_favorite_user_ids(self):
        return [(6, 0, self.env['hr.department'].search([]).mapped('id'))]

    def _get_default_socio_cat_levels_ids(self):
        return [(6, 0, self.env['hr.socioprofessional.categories.levels'].search([]).mapped('id'))]

    def _get_default_employee_ids(self):
        return [(6, 0, self.env['hr.employee'].search([]).mapped('id'))]

    date_start = fields.Date(related='payslip_run_id.date_start', string="Date From")
    date_end = fields.Date(related='payslip_run_id.date_end', string="Date End")
    payslip_run_id = fields.Many2one('hr.payslip.run', string='Payslip Batches', required=True)
    department_ids = fields.Many2many('hr.department', string='HR Departments', default=_get_default_favorite_user_ids)
    employee_ids = fields.Many2many('hr.employee', string='HR Employee',
                                    relation='create_payroll_journal_hr_employee_smartest',
                                    compute='_compute_employee_ids')
    csp_ids = fields.Many2many('hr.socioprofessional.categories.levels', string='HR Departments',
                               relation='create_payroll_journal_hr_socio_cat_levels_rel',
                               default=_get_default_socio_cat_levels_ids)

    report = fields.Binary('', readonly=True)
    report_name = fields.Char('Report Name', readonly=True)

    livre_per = fields.Selection([
        ('department', 'Departement'),
        ('employee', 'Employee'),
        ('CSP', 'CSP'),
    ], required=True)

    livre_type = fields.Selection([
        ('horiz', 'Horizontal'),
        ('vert', 'Vertical'),
    ], required=True)

    @api.depends('payslip_run_id')
    def _compute_employee_ids(self):
        for rec in self:
            rec.employee_ids = rec.payslip_run_id.slip_ids.mapped('employee_id')

    def print_report(self):
        salary_rules = self.env.ref('smartest_hr_payroll.structure_base_dz').rule_ids.filtered(
            lambda p: p.appears_on_payslip and p.sequence not in [501, 1500])
        data = []

        rule_data2 = []
        total_cotisation = 0
        total_irg = 0
        cotisation = [0 for department in self.department_ids]
        irg = [0 for department in self.department_ids]
        for rule in salary_rules:
            rule_name = 'Retenue IRG absence' if 720 == rule.sequence else rule.name
            rule_data = [rule.sequence, rule_name]
            total = 0
            index = 0
            for department in self.department_ids:
                slip_lines = self.payslip_run_id.slip_ids.filtered(lambda p: p.state in ['done', 'verify',
                                                                                         'paid'] and p.employee_id.department_id and department.id == p.employee_id.department_id.id).mapped(
                    'line_ids').filtered(lambda p: p.salary_rule_id.id == rule.id).mapped('total')
                total += round(abs(sum(slip_lines)), 2)
                if rule.sequence in [505, 507] and total_cotisation >= 0:
                    total_cotisation += round(abs(sum(slip_lines)), 2)
                    cotisation[index] += round(abs(sum(slip_lines)), 2)
                rule_data.append(round(abs(sum(slip_lines)), 2))
                index += 1

            if total > 0:
                if rule.sequence > 507 and total_cotisation >= 0:
                    data.append((['\t', 'Total Cotisation salariale'] + cotisation + [total_cotisation]))
                    total_cotisation = -1
                rule_data.append(round(total, 2))
                data.append(rule_data)

        salary_rules = self.env.ref('smartest_hr_payroll.structure_base_dz').rule_ids.filtered(
            lambda p: p.sequence in [500, 501, 501, 507, 506, 905, 1500])
        data2 = []
        total_employer_contribution = 0
        employer_contribution = [0 for department in self.department_ids]
        for rule in salary_rules:
            rule_name2 = 'Salaire de poste' if 500 == rule.sequence else rule.name
            rule_data2 = [rule.sequence, rule_name2]
            total2 = 0
            index3 = 0
            for department in self.department_ids:
                if rule.sequence == 905:
                    slip_lines = self.payslip_run_id.slip_ids.filtered(lambda p: p.state in ['done', 'verify',
                                                                                             'paid'] and p.employee_id.department_id and department.id == p.employee_id.department_id.id).mapped(
                        'line_ids').filtered(lambda p: p.salary_rule_id.id == rule.id).mapped('amount')
                    total2 += round(abs(sum(slip_lines)), 2)
                else:
                    slip_lines = self.payslip_run_id.slip_ids.filtered(lambda
                                                                           p: p.state in ['done', 'verify',
                                                                                          'paid'] and p.employee_id.department_id and department.id == p.employee_id.department_id.id).mapped(
                        'line_ids').filtered(lambda p: p.salary_rule_id.id == rule.id).mapped('total')
                    total2 += round(abs(sum(slip_lines)), 2)

                if rule.sequence in [506] and total_employer_contribution >= 0:
                    total_employer_contribution += round(abs(sum(slip_lines)), 2)
                    employer_contribution[index3] += round(abs(sum(slip_lines)), 2)
                rule_data2.append(round(abs(sum(slip_lines)), 2))
                index3 += 1
            if total2 > 0:

                if rule.sequence > 506 and total_employer_contribution >= 0:
                    data.append(
                        (['\t', 'Cotisations patronales'] + employer_contribution + [total_employer_contribution]))
                    total_employer_contribution = -1
                rule_data2.append(round(total2, 2))

                data2.append(rule_data2)
        #   Extraction of net taxable**************************************************************

        return self.env.ref('smartest_hr_payroll.payroll_journal_report').report_action(self, data={
            'rules': data,
            'rules2': data2,
            'departments': self.department_ids.mapped('name'),
            'company': self.env.user.company_id.name,
            'date_now': datetime.now().strftime("%d/%m/%Y"),
            'hour_now': datetime.now().strftime("%H:%M"),
            'date_start': self.payslip_run_id.date_start.strftime("%d/%m/%Y"),
            'date_end': self.payslip_run_id.date_end.strftime("%d/%m/%Y"),
            'nbr_employee': len(self.payslip_run_id.slip_ids.filtered(lambda p: p.state in ['done', 'verify', 'paid']))
        })

    # This part was requested from RAYANOX
    def print_report_csp(self):
        salary_rules = self.env.ref('smartest_hr_payroll.structure_base_dz').rule_ids.filtered(
            lambda p: p.appears_on_payslip and p.sequence not in [501, 1500])
        data = []

        rule_data2 = []
        total_cotisation = 0
        total_irg = 0
        cotisation = [0 for csp in self.csp_ids]
        irg = [0 for csp in self.csp_ids]
        for rule in salary_rules:
            rule_name = 'Retenue IRG absence' if 720 == rule.sequence else rule.name
            rule_data = [rule.sequence, rule_name]
            total = 0
            index = 0
            for csp in self.csp_ids:
                slip_lines = self.payslip_run_id.slip_ids.filtered(
                    lambda p: p.state in ['done', 'verify', 'paid'] and p.csp_cat and csp.id == p.csp_cat.id).mapped(
                    'line_ids').filtered(lambda p: p.salary_rule_id.id == rule.id).mapped('total')
                total += round(abs(sum(slip_lines)), 2)
                if rule.sequence in [501, 507] and total_cotisation >= 0:
                    total_cotisation += round(abs(sum(slip_lines)), 2)
                    cotisation[index] += round(abs(sum(slip_lines)), 2)
                rule_data.append(round(abs(sum(slip_lines)), 2))
                index += 1

            if total > 0:
                if rule.sequence > 507 and total_cotisation >= 0:
                    data.append((['\t', 'Total Cotisation salariale'] + cotisation + [total_cotisation]))
                    total_cotisation = -1
                rule_data.append(round(total, 2))
                data.append(rule_data)

        salary_rules = self.env.ref('smartest_hr_payroll.structure_base_dz').rule_ids.filtered(
            lambda p: p.sequence in [500, 501, 501, 507, 506, 905, 1500])
        data2 = []
        total_employer_contribution = 0
        employer_contribution = [0 for department in self.department_ids]
        for rule in salary_rules:
            rule_name2 = 'Salaire de poste' if 500 == rule.sequence else rule.name
            rule_data2 = [rule.sequence, rule_name2]
            total2 = 0
            index3 = 0
            for csp in self.csp_ids:
                if rule.sequence == 905:
                    slip_lines = self.payslip_run_id.slip_ids.filtered(lambda p: p.state in ['done', 'verify',
                                                                                             'paid'] and p.csp_cat and csp.id == p.csp_cat.id).mapped(
                        'line_ids').filtered(lambda p: p.salary_rule_id.id == rule.id).mapped('amount')
                    total2 += round(abs(sum(slip_lines)), 2)
                else:
                    slip_lines = self.payslip_run_id.slip_ids.filtered(lambda
                                                                           p: p.state in ['done', 'verify',
                                                                                          'paid'] and p.csp_cat and csp.id == p.csp_cat.id).mapped(
                        'line_ids').filtered(lambda p: p.salary_rule_id.id == rule.id).mapped('total')
                    total2 += round(abs(sum(slip_lines)), 2)

                if rule.sequence in [506] and total_employer_contribution >= 0:
                    total_employer_contribution += round(abs(sum(slip_lines)), 2)
                    employer_contribution[index3] += round(abs(sum(slip_lines)), 2)
                rule_data2.append(round(abs(sum(slip_lines)), 2))
                index3 += 1
            if total2 > 0:

                if rule.sequence > 506 and total_employer_contribution >= 0:
                    data.append(
                        (['\t', 'Cotisations patronales'] + employer_contribution + [total_employer_contribution]))
                    total_employer_contribution = -1
                rule_data2.append(round(total2, 2))

                data2.append(rule_data2)
        #         Extraction of net taxable**************************************************************

        return self.env.ref('smartest_hr_payroll.payroll_journal_report_csp').report_action(self, data={
            'rules': data,
            'rules2': data2,
            'departments': self.csp_ids.mapped('name'),
            'company': self.env.user.company_id.name,
            'date_now': datetime.now().strftime("%d/%m/%Y"),
            'hour_now': datetime.now().strftime("%H:%M"),
            'date_start': self.payslip_run_id.date_start.strftime("%d/%m/%Y"),
            'date_end': self.payslip_run_id.date_end.strftime("%d/%m/%Y"),
            'nbr_employee': len(self.payslip_run_id.slip_ids.filtered(lambda p: p.state in ['done', 'verify', 'paid']))
        })

    def print_report_emp(self):
        salary_rules = self.env.ref('smartest_hr_payroll.structure_base_dz').rule_ids.filtered(
            lambda p: p.appears_on_payslip and p.sequence not in [501, 1500])
        data = []

        rule_data2 = []
        total_cotisation = 0
        total_irg = 0
        cotisation = [0 for employee in self.employee_ids]
        irg = [0 for employee in self.employee_ids]
        for rule in salary_rules:
            rule_name = 'Retenue IRG absence' if 720 == rule.sequence else rule.name
            rule_data = [rule.sequence, rule_name]
            total = 0
            index = 0
            for employee in self.employee_ids:
                slip_lines = self.payslip_run_id.slip_ids.filtered(lambda
                                                                       p: p.employee_id and employee.id == p.employee_id.id).mapped(
                    'line_ids').filtered(lambda p: p.salary_rule_id.id == rule.id).mapped('total')
                total += round(abs(sum(slip_lines)), 2)
                if rule.sequence in [507, 505] and total_cotisation >= 0:
                    total_cotisation += round(abs(sum(slip_lines)), 2)
                    cotisation[index] += round(abs(sum(slip_lines)), 2)
                rule_data.append(round(abs(sum(slip_lines)), 2))
                index += 1

            if total > 0:
                if rule.sequence > 507 and total_cotisation >= 0:
                    data.append((['\t', 'Total Cotisation salariale'] + cotisation + [total_cotisation]))
                    total_cotisation = -1
                rule_data.append(round(total, 2))
                data.append(rule_data)

        salary_rules = self.env.ref('smartest_hr_payroll.structure_base_dz').rule_ids.filtered(
            lambda p: p.sequence in [500, 501, 501, 507,506, 905, 1500])
        data2 = []
        total_employer_contribution = 0
        employer_contribution = [0 for employee in self.employee_ids]
        for rule in salary_rules:
            rule_name2 = 'Salaire de poste' if 500 == rule.sequence else rule.name
            rule_data2 = [rule.sequence, rule_name2]
            total2 = 0
            index3 = 0
            for employee in self.employee_ids:
                if rule.sequence == 905:
                    slip_lines = self.payslip_run_id.slip_ids.filtered(lambda
                                                                           p: p.employee_id and employee.id == p.employee_id.id).mapped(
                        'line_ids').filtered(lambda p: p.salary_rule_id.id == rule.id).mapped('amount')
                    total2 += round(abs(sum(slip_lines)), 2)
                else:
                    slip_lines = self.payslip_run_id.slip_ids.filtered(lambda
                                                                           p: p.employee_id and employee.id == p.employee_id.id).mapped(
                        'line_ids').filtered(lambda p: p.salary_rule_id.id == rule.id).mapped('total')
                    total2 += round(abs(sum(slip_lines)), 2)

                if rule.sequence in [506] and total_employer_contribution >= 0:
                    total_employer_contribution += round(abs(sum(slip_lines)), 2)
                    employer_contribution[index3] += round(abs(sum(slip_lines)), 2)
                if rule.sequence in [1500]:
                    rule_data2.append(round((sum(slip_lines)), 2))
                else:
                    rule_data2.append(round(abs(sum(slip_lines)), 2))
                index3 += 1
            if total2 > 0:

                if rule.sequence > 506 and total_employer_contribution >= 0:
                    data.append(
                        (['\t', 'Cotisations patronales'] + employer_contribution + [total_employer_contribution]))
                    total_employer_contribution = -1
                rule_data2.append(round(total2, 2))

                data2.append(rule_data2)
        #         Extraction of net taxable**************************************************************

        return self.env.ref('smartest_hr_payroll.payroll_journal_report').report_action(self, data={
            'rules': data,
            'rules2': data2,
            'departments': self.employee_ids.mapped('name'),
            'employee_mat': self.employee_ids.mapped('registration_number'),
            'company': self.env.user.company_id.name,
            'date_now': datetime.now().strftime("%d/%m/%Y"),
            'hour_now': datetime.now().strftime("%H:%M"),
            'date_start': self.payslip_run_id.date_start.strftime("%d/%m/%Y"),
            'date_end': self.payslip_run_id.date_end.strftime("%d/%m/%Y"),
            'nbr_employee': len(self.payslip_run_id.slip_ids.filtered(lambda p: p.state == 'done'))
        })

    def print_report_horiz(self):
        path = self.env.ref('smartest_base.smartest_files_path').sudo().value
        ind = 1
        if self.livre_per == 'department':
            var = self.print_report()
        elif self.livre_per == 'CSP':
            var = self.print_report_csp()
        else:
            var = self.print_report_emp()
            ind = 2

        data = var['data']
        rule = []
        value = []
        workbook = Workbook()
        redFill = PatternFill(start_color='90a4ae',
                              end_color='90a4ae',
                              fill_type='solid',
                              )
        orangeFill = PatternFill(start_color='FFA500',
                              end_color='FFA500',
                              fill_type='solid',
                              )

        yellowFill = PatternFill(start_color='FFFF00',
                                 end_color='FFFF00',
                                 fill_type='solid',
                                 )

        top = Side(border_style='thin', color="000000")
        bottom = Side(border_style='thin', color="000000")
        left = Side(border_style='thin', color="000000")
        right = Side(border_style='thin', color="000000")
        border = Border(top=top, bottom=bottom, left=left, right=right)
        i = ind
        sheet = workbook.active
        sheet.merge_cells(start_row=1, start_column=4, end_row=3, end_column=7)
        sheet.merge_cells(start_row=4, start_column=4, end_row=5, end_column=7)
        sheet.merge_cells(start_row=1, start_column=8, end_row=3, end_column=11)
        sheet.merge_cells(start_row=4, start_column=8, end_row=5, end_column=11)
        sheet.merge_cells(start_row=6, start_column=4, end_row=6, end_column=11)

        sheet.merge_cells(start_row=8, start_column=4, end_row=8, end_column=7)

        for x in range(4, 8):
            sheet.cell(row=8, column=x).border = border

        sheet["D1"].fill = redFill
        sheet["H1"].fill = redFill
        sheet["D4"].fill = redFill
        sheet["H4"].fill = redFill
        sheet["D6"].fill = redFill
        sheet["D1"].value = 'Date :' + str(data['date_now'])
        sheet["D4"].value = 'Heure :' + str(data['hour_now'])
        sheet["H4"].value = str(data['date_start']) + ' - ' + str(data['date_end'])
        sheet["H1"].value = 'Livre de paie'
        sheet["D8"].value = 'Société : ' + str(data['company'])
        sheet["H1"].font = Font(size=20)
        sheet.cell(row=10, column=i).value = 'Rebrique'
        sheet.cell(row=10, column=i).fill = redFill
        sheet.cell(row=10, column=i).border = border
        for rec in data['rules']:
            i += 1

            sheet.cell(row=10, column=i).value = str(rec[0]) + ' ' + rec[1]
            sheet.cell(row=10, column=i).fill = redFill
            sheet.cell(row=10, column=i).border = border
            if 'Total' in rec[1]:
                sheet.cell(row=10, column=i).fill = yellowFill
                j = 11
                for dep in data['departments']:
                    sheet.cell(row=j, column=i).value = rec[j - 9]
                    sheet.cell(row=j, column=i).fill = yellowFill
                    sheet.cell(row=j, column=i).border = border
                    # sheet.cell(row=i-1, column=j).value = rec[j+1]
                    j += 1

                sheet.cell(row=j, column=i).value = rec[j - 9]
                sheet.cell(row=j, column=i).fill = yellowFill
                sheet.cell(row=j, column=i).border = border

            rule = data['rules']
            j = 11
            for dep in data['departments']:
                sheet.cell(row=j, column=i).value = rec[j - 9]
                sheet.cell(row=j, column=i).border = border
                # sheet.cell(row=i-1, column=j).value = rec[j+1]
                j += 1
            sheet.cell(row=j, column=i).value = rec[j - 9]
            sheet.cell(row=j, column=i).border = border

        if self.livre_per == 'employee':
            j = 11
            for dep in data['employee_mat']:
                sheet.cell(row=j, column=1).value = dep
                sheet.cell(row=j, column=1).fill = redFill
                sheet.cell(row=j, column=1).border = border
                j += 1

        j = 11
        for dep in data['departments']:
            sheet.cell(row=j, column=ind).value = dep
            sheet.cell(row=j, column=ind).fill = redFill
            sheet.cell(row=j, column=ind).border = border
            j += 1

        sheet.cell(row=j, column=ind).value = 'Total'
        sheet.cell(row=j + 1, column=ind).value = 'Nombre de salariés'
        sheet.cell(row=j, column=ind).fill = redFill
        sheet.cell(row=j, column=ind).border = border
        sheet.cell(row=j + 1, column=ind).fill = redFill
        sheet.cell(row=j + 1, column=ind).border = border
        sheet.merge_cells(start_row=j + 1, start_column=ind+1, end_row=j + 1, end_column=len(data['rules']))
        sheet.cell(row=j + 1, column=2).value = data['nbr_employee']

        for indice in range(len(data['rules']) - 1):
            sheet.cell(row=j + 1, column=indice + 2).border = border
        for rec in data['rules2']:
            if rec[0] not in [501, 501, 507]:
                i += 1
                if rec[0] in [905]:
                    sheet.cell(row=10, column=i).value = 'BASE imposable IRG'
                    sheet.cell(row=10, column=i).fill = redFill
                    sheet.cell(row=10, column=i).border = border
                else:
                    sheet.cell(row=10, column=i).value = rec[1]
                    sheet.cell(row=10, column=i).fill = redFill
                    sheet.cell(row=10, column=i).border = border
                j = 11
                for dep in data['departments']:
                    sheet.cell(row=j, column=i).value = rec[j - 9]
                    if rec[j-9] < 0:
                        sheet.cell(row=j, column=i).fill = orangeFill
                    sheet.cell(row=j, column=i).border = border
                    j += 1
                sheet.cell(row=j, column=i).value = rec[j - 9]
                sheet.cell(row=j, column=i).border = border

        if self.livre_per == 'department':
            workbook.save(filename=path + "Livre de paie par Departements.xlsx")
            file = open(path + "Livre de paie par Departements.xlsx", "rb")
            out = file.read()
            file.close()
            self.report = base64.b64encode(out)
            os.remove(path + "Livre de paie par Departements.xlsx")
            self.report_name = "Livre de paie par Departements(" + str(data['date_start']) + ' - ' + str(
                data['date_end']) + ").xlsx"
        if self.livre_per == 'CSP':
            workbook.save(filename=path + "Livre de paie par CSP.xlsx")
            file = open(path + "Livre de paie par CSP.xlsx", "rb")
            out = file.read()
            file.close()
            self.report = base64.b64encode(out)
            os.remove(path + "Livre de paie par CSP.xlsx")
            self.report_name = "Livre de paie par CSP(" + str(data['date_start']) + ' - ' + str(
                data['date_end']) + ")..xlsx"
        if self.livre_per == 'employee':
            workbook.save(filename=path + "Livre de paie par employee.xlsx")
            file = open(path + "Livre de paie par employee.xlsx", "rb")
            out = file.read()
            file.close()
            self.report = base64.b64encode(out)
            os.remove(path + "Livre de paie par employee.xlsx")
            self.report_name = "Livre de paie par employee(" + str(data['date_start']) + ' - ' + str(
                data['date_end']) + ")..xlsx"

        return {
            'name': ('Créer le livre de paie'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'create.payroll.journal',
            'target': 'new',
            'res_id': self.id,
        }
