from odoo import _, api, fields, models
from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl import drawing as Img
from odoo.exceptions import UserError
import base64, datetime
import os
import shutil
from odoo.exceptions import ValidationError



class SmartestGfiftyAccountMoveLine(models.Model):
    _name='g_cinquante'
    _inherit = ['mail.thread']


    name = fields.Char('name', readonly=1)
    ref = fields.Char('Refrence')
    state = fields.Selection([('draft', 'Draft'),('ongoing', 'On going'), ('approved', 'Approved'),('done', 'Done')],string = 'state', default='draft')
    date_from = fields.Date('Date from')
    date_to = fields.Date('Date to')
    raw_amount = fields.Float("Chiffre d'affaires brut")
    raw_amount_column = fields.Integer('Colonne total brut' , default=6)
    raw_amount_row = fields.Integer('Ligne total brut', default=13)
    g50_lines = fields.One2many('g_cinquante.line','g_fifty',string='Lines', domain=[('is_purchase_invoice', '!=', 'True')])
    g50_lines1 = fields.One2many('g_cinquante.line','g_fifty',string='Lines1', domain=[('is_purchase_invoice', '=', 'True')])
    attachment = fields.Binary('Download File')
    total_imposable = fields.Float("Total Imposable")
    total_imposable_column = fields.Integer('Colonne Total Imposable', default=8)
    total_imposable_row = fields.Integer('Ligne Total Imposable', default=20)
    total_exonere = fields.Float('Total Exonéré')
    total_exonere_column = fields.Integer('Colonne Total Exonéré', default=7)
    total_exonere_row = fields.Integer('Ligne Total Exonéré', default=20)
    tva_purchase_inves_1st_row = fields.Integer('1st Ligne TVA Achat Investissement', default=19)
    tva_purchase_stck_1st_row = fields.Integer('1st Ligne TVA Achat stock', default=31)
    company_id = fields.Many2one('res.company',string="Company",required=True,default=lambda self:self.env.company)
    purchase_inves_journal_id = fields.Many2one('account.journal', string='Journal purchase Investissment')
    purchase_stck_journal_id = fields.Many2one('account.journal', string='Journal purchase Stock')
    payslip_run_id = fields.Many2one('hr.payslip.run', string='Payslip Batches')
    p2_impot_ids = fields.One2many('g_cinquante.impot_line','p2_id',string="P2")
    doc_path = fields.Char(default="/home/odoo/attachement/")
    addons_path = fields.Char(default="/home/odoo/repositories/")

    def _delete_attachment_and_dic_from_os(self):
        for this in self:
            try :
                shutil.rmtree(this.doc_path)
            except :
                pass

    def _create_saving_directory(self):
        for this in self:
            try:
                os.makedirs(this.doc_path, exist_ok = True)
            except OSError as error:
                raise ValidationError(
                    _(error)
                )


    def get_p2_values(self):
        for this in self:
            p2_lines = self.env['g_cinquante.impot_line'].search([])
            for line in p2_lines:
                line.p2_id = this.id
                line.compute_account_id()



    def test_irg(self):
        """get irg from functio that compute livre de paie"""
        for this in self:
            paylips_move_lines = self.env['account.move.line'].search([('move_id.ref','=',this.payslip_run_id.name)])
            for line in this.p2_impot_ids.filtered(lambda this: this.has_payslips_relation):
                for move_line in paylips_move_lines:
                    if line.account_id == move_line.account_id:
                        line.total = move_line.debit + move_line.credit
            for line in this.p2_impot_ids.filtered(lambda this: not this.has_payslips_relation):
                line_moves = self.env['account.move.line'].search([('account_id','=',line.account_id.id),('date','>=',this.date_from),('date','<=',this.date_to),('move_id.state','=','posted')])
                line.total = sum(line_moves.mapped('debit')) + sum (line_moves.mapped('credit'))



    def draft_g50(self):
        for record in self :
            record.state = 'draft'
            self.ensure_one()
            self.env.cr.execute(
                """
                delete from g_cinquante_line where (g_fifty = %s)
                """%record.id)
            return

    def compute_g50_(self):
        """this fct compute necessary values for g50"""
        for this in self:
            this.test_irg()
            total_imposable = 0
            total_exonéré = 0
            for line in this.g50_lines:
                if line.invoice_totat_taxe == 0 :
                    total_exonéré += line.payment_id.amount_currency
                    print(line.payment_id.amount_currency)
                elif line.invoice_totat_taxe != 0 :
                    total_imposable += line.payment_id.amount_currency
            this.total_exonere = total_exonéré
            this.total_imposable = total_imposable
            this.raw_amount = total_exonéré + total_imposable
            this.state = 'approved'


    def prepare_g50_lines_(self):
        """this fonction prepare the payments move lines and its value"""
        for this in self:
            this.get_p2_values()
            move_line = this.env['account.move.line'].search([('account_id.code','like','512'),('date','>=',this.date_from),('date','<=',this.date_to),('move_id.state','=','posted')])
            purchase_invoice_move_ = this.env['account.move'].search([('state', '=', 'posted'),('move_type','=','in_invoice'),('date','>=',this.date_from),('date','<=',this.date_to)])
            for move in move_line:
                invoice_move_ = this.env['account.move'].search([('state','=','posted'),('name','=',move.ref)])
                this.env['g_cinquante.line'].create({
                        'g_fifty': this.id,
                        'payment_id': move.id,
                        'invoice_id' : invoice_move_.id,
                        'invoice_totat_taxe' : invoice_move_.amount_tax,
                        'is_purchase_invoice': False,
                    })
            for invoice in purchase_invoice_move_:
                this.env['g_cinquante.line'].create({
                        'g_fifty': this.id,
                        'purchase_invoice_id' : invoice.id,
                        'is_purchase_invoice' : True,
                    })
            this.state = 'ongoing'


    def print_g50_(self):
        """function that use the G50 report as template and save a it as new"""
        for rec in self:
            rec._create_saving_directory()
            def add_images_to_report(rec):
                """this fonction paste the pictures in the right place in the report"""

                image_G_50_month_year = Img.image.Image(
                    rec.addons_path+'smartest-odoo-app/smartest_g50/static/img/G_50_month_year.png')
                image_G_50_month_year.width = 170; image_G_50_month_year.height = 30
                image_G_50_month_year.anchor = 'D1'
                image_G_50_code_activite = Img.image.Image(
                    rec.addons_path+'smartest-odoo-app/smartest_g50/static/img/G_50_code_activite.png')
                image_G_50_code_activite.width = 180; image_G_50_code_activite.height = 70
                image_G_50_code_activite.anchor = 'I7'
                image_G_50_g_j = Img.image.Image(
                    rec.addons_path+'smartest-odoo-app/smartest_g50/static/img/G_50_f_j.png')
                image_G_50_g_j.width = 180; image_G_50_g_j.height = 80
                image_G_50_g_j.anchor = 'D6'
                image_G50_attention = Img.image.Image(
                    rec.addons_path+'smartest-odoo-app/smartest_g50/static/img/G50_attention.png')
                image_G50_attention.anchor = 'I2'
                image_G50_attention.width = 200; image_G50_attention.height = 120
                image_G_50_trimestre = Img.image.Image(
                    rec.addons_path+'smartest-odoo-app/smartest_g50/static/img/G_50_trimestre.png')
                image_G_50_trimestre.width = 170; image_G_50_trimestre.height = 50
                image_G_50_trimestre.anchor = 'D3'
                sheet1.add_image(image_G_50_month_year)
                sheet1.add_image(image_G_50_code_activite)
                sheet1.add_image(image_G_50_g_j)
                sheet1.add_image(image_G50_attention)
                sheet1.add_image(image_G_50_trimestre)

            workbook = load_workbook(rec.addons_path+"smartest-odoo-app/smartest_g50/static/file/Copie de G50 SMARTEST 03-2021.xlsx")
            sheets = workbook.sheetnames

            # page 1
            sheet1 = workbook[sheets[1]]
            sheet1.cell(row=rec.raw_amount_row, column=rec.raw_amount_column).value = rec.raw_amount

            # page 2
            sheet2 = workbook[sheets[2]]
            for p2_line in rec.p2_impot_ids:
                sheet2.cell(row=p2_line.row, column=p2_line.column).value = p2_line.total

            # page 3
            sheet3 = workbook[sheets[3]]
            sheet3.cell(row=rec.total_exonere_row, column=rec.total_exonere_column).value = rec.total_exonere
            sheet3.cell(row=rec.total_imposable_row, column=rec.total_imposable_column).value = rec.total_imposable

            #page tva
            sheet4 = workbook[sheets[4]]
            row_inv = rec.tva_purchase_inves_1st_row
            row_stk = rec.tva_purchase_stck_1st_row
            for line in rec.g50_lines1 :
                if line.purchase_invoice_tva_inves != 0:
                    row_inv += 1
                    sheet4.cell(row=row_inv, column=2).value = line.purchase_invoice_partner_id.name
                    sheet4.cell(row=row_inv, column=3).value = line.purchase_invoice_partner_id.street
                    sheet4.cell(row=row_inv, column=4).value = line.purchase_invoice_partner_id.fiscal_identification
                    sheet4.cell(row=row_inv, column=5).value = line.purchase_invoice_partner_id.commercial_register
                    sheet4.cell(row=row_inv, column=6).value = line.purchase_invoice_partner_id.taxation
                    sheet4.cell(row=row_inv, column=7).value = line.purchase_invoice_id.name
                    sheet4.cell(row=row_inv, column=8).value = line.purchase_invoice_date
                    sheet4.cell(row=row_inv, column=9).value = line.purchase_invoice_amount_ht
                    sheet4.cell(row=row_inv, column=10).value = line.purchase_invoice_tva_inves
                    sheet4.cell(row=row_inv, column=11).value = line.purchase_invoice_tva_stk
                    sheet4.cell(row=row_inv, column=12).value = line.purchase_invoice_tva_inves
                elif line.purchase_invoice_tva_stk != 0:
                    row_stk += 1
                    sheet4.cell(row=row_stk, column=2).value = line.purchase_invoice_partner_id.name
                    sheet4.cell(row=row_stk, column=3).value = line.purchase_invoice_partner_id.street
                    sheet4.cell(row=row_stk, column=4).value = line.purchase_invoice_partner_id.fiscal_identification
                    sheet4.cell(row=row_stk, column=5).value = line.purchase_invoice_partner_id.commercial_register
                    sheet4.cell(row=row_stk, column=6).value = line.purchase_invoice_partner_id.taxation
                    sheet4.cell(row=row_stk, column=7).value = line.purchase_invoice_id.name
                    sheet4.cell(row=row_stk, column=8).value = line.purchase_invoice_date
                    sheet4.cell(row=row_stk, column=9).value = line.purchase_invoice_amount_ht
                    sheet4.cell(row=row_stk, column=10).value = line.purchase_invoice_tva_inves
                    sheet4.cell(row=row_stk, column=11).value = line.purchase_invoice_tva_stk
                    sheet4.cell(row=row_stk, column=12).value = line.purchase_invoice_tva_stk

            #save file
            add_images_to_report(rec)
            file_name = 'G50_report'+ rec.name + '.xlsx'
            workbook.save(rec.doc_path+file_name)
            rec.attachment = base64.b64encode(open(rec.doc_path+file_name, 'rb').read())
            rec.state = 'done'
            rec._delete_attachment_and_dic_from_os()
            # self.create_attachement(file_name)

    # def create_attachement(self,file_name):
    #     file = open('/home/smartest/Bureau/imp/'+file_name, 'rb')
    #     self.env['ir.attachment'].sudo().create({
    #         'name'        : file_name,
    #         'res_model'   : 'g_cinquante',
    #         'res_id'      : self.id,
    #         'type'        : 'binary',
    #         'mimetype'    : 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    #         'datas'       : base64.b64encode(file.read()),
    #         # 'description' : 'files',
    #         })
    #
    #     return {
    #         'type': 'ir.actions.client',
    #         'tag': 'reload',
    #         'target' : self.id,
    #     }

    @api.model
    def create(self, vals):
        if not vals.get('name') == _('New'):
            vals['name'] = self.env["ir.sequence"].next_by_code("g_cinquante") or _('/')

        return super(SmartestGfiftyAccountMoveLine, self).create(vals)

class SmartestGfiftyLine(models.Model):
    _name='g_cinquante.line'

    payment_id = fields.Many2one('account.move.line',string='Payment move line', ondelete='set null', domain="[('account_id.code','like','512')]")
    invoice_id = fields.Many2one('account.move', string='Invoice move', ondelete='set null', domain="[('state','=','posted'),('move_type','=','out_invoice')]")
    purchase_invoice_id = fields.Many2one('account.move', string='Purchase Invoice', ondelete='set null', domain="[('state','=','posted'),('move_type','=','in_invoice')]")
    purchase_invoice_partner_id = fields.Many2one('res.partner',related="purchase_invoice_id.partner_id")
    purchase_invoice_date = fields.Date(related="purchase_invoice_id.date")
    purchase_invoice_amount_ht = fields.Float(string="total untaxed", compute="_onchange_invoice_total_tax")
    purchase_invoice_tva_inves = fields.Float(string="TVA Inves")
    purchase_invoice_tva_stk = fields.Float(string="TVA Stock")
    invoice_totat_taxe = fields.Float(string='Total Taxe')
    is_purchase_invoice = fields.Boolean(default=False)
    g_fifty = fields.Many2one('g_cinquante')
    company_id = fields.Many2one('res.company',string="Company",required=True,default=lambda self:self.env.company)


    @api.depends('invoice_id','purchase_invoice_id')
    def _onchange_invoice_total_tax(self):
        for line in self:
            line.invoice_totat_taxe = line.invoice_id.amount_tax
            line.purchase_invoice_amount_ht = line.purchase_invoice_id.amount_untaxed
            if line.purchase_invoice_id.journal_id.code == 'inv':
                line.purchase_invoice_tva_inves = line.purchase_invoice_id.amount_tax
            elif line.purchase_invoice_id.journal_id.code == 'FACT':
                line.purchase_invoice_tva_stk = line.purchase_invoice_id.amount_tax
            else:
                line.purchase_invoice_tva_inves = 0
                line.purchase_invoice_tva_stk = 0

class SmartestGfiftyLineImpot(models.Model):
    _name='g_cinquante.impot_line'

    p2_id = fields.Many2one('g_cinquante')
    name = fields.Char()
    account_id = fields.Many2one('account.account',string="Account")
    account_code = fields.Char(string="Char")
    column = fields.Integer('Colonne')
    row = fields.Integer('Ligne')
    total = fields.Integer('Total')
    has_payslips_relation = fields.Boolean(default=False)

    def compute_account_id(self):
        for line in self:
            line.account_id = self.env['account.account'].search([('code','=',line.account_code)])
            line.total = 0

